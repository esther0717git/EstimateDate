# Modified Visitor List Cleaner to preserve all sheets

import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# --- Streamlit setup ---
st.set_page_config(page_title="Visitor List Cleaner", layout="wide")
st.title("\U0001F1F8\U0001F1EC CLARITY GATE – VISITOR DATA CLEANING & VALIDATION \U0001FAE7")

with open("sample_template.xlsx", "rb") as f:
    sample_bytes = f.read()
st.download_button(
    label="🌟 Download Sample Template",
    data=sample_bytes,
    file_name="sample_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.info("""
**Data Integrity Is Our Foundation**  
At every step—from file upload to final report—we enforce strict validation to guarantee your visitor data is accurate, complete, and compliant.  
Maintaining integrity not only expedites gate clearance, it protects our facilities and ensures we meet all regulatory requirements.
""")

with st.expander("Why is Data Integrity Important?"):
    st.write("""
    - **Accuracy**: Correct visitor details reduce clearance delays.  
    - **Security**: Reliable ID checks prevent unauthorized access.  
    - **Compliance**: Audit-ready records ensure regulatory adherence.  
    - **Efficiency**: Trustworthy data powers faster reporting and analytics.
    """)

st.markdown("### ⚠️ **Please ensure your spreadsheet has no missing or malformed fields.**")
uploaded = st.file_uploader("📁 Upload your Excel file", type=["xlsx"])

# --- Date Estimation ---
now = datetime.now(ZoneInfo("Asia/Singapore"))
st.markdown("### 🗓️ Estimate Clearance Date 🍍")
formatted_now = now.strftime("%A %d %B, %I:%M%p").lstrip("0")
st.write("**Today is:**", formatted_now)

if st.button("▶️ Calculate Estimated Delivery"):
    if now.time() >= datetime.strptime("15:00", "%H:%M").time():
        effective_submission_date = now.date() + timedelta(days=1)
    else:
        effective_submission_date = now.date()

    while effective_submission_date.weekday() >= 5:
        effective_submission_date += timedelta(days=1)

    working_days_count = 0
    estimated_date = effective_submission_date
    while working_days_count < 2:
        estimated_date += timedelta(days=1)
        if estimated_date.weekday() < 5:
            working_days_count += 1

    clearance_date = estimated_date
    while clearance_date.weekday() >= 5:
        clearance_date += timedelta(days=1)

    formatted = f"{clearance_date:%A} {clearance_date.day} {clearance_date:%B}"
    st.success(f"✓ Earliest clearance: **{formatted}**")

# --- Cleaning Function ---
# (keep your existing clean_data() function as-is)

# --- Excel Writer that preserves all sheets ---
def generate_full_workbook(all_sheets: dict) -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            if sheet_name != "Visitor List":
                continue

            # Apply formatting only to Visitor List tab
            header_fill  = PatternFill("solid", fgColor="94B455")
            warning_fill = PatternFill("solid", fgColor="DA9694")
            border       = Border(*[Side("thin")] * 4)
            center       = Alignment("center", "center")
            normal_font  = Font(name="Calibri", size=9)
            bold_font    = Font(name="Calibri", size=9, bold=True)

            for row in ws.iter_rows():
                for cell in row:
                    cell.border    = border
                    cell.alignment = center
                    cell.font      = normal_font

            for col in range(1, ws.max_column + 1):
                h = ws[f"{get_column_letter(col)}1"]
                h.fill = header_fill
                h.font = bold_font
            ws.freeze_panes = "B2"

            errors = 0
            seen = {}

            for r in range(2, ws.max_row + 1):
                idt = str(ws[f"G{r}"].value).strip().upper()
                nat = str(ws[f"J{r}"].value).strip().title()
                pr  = str(ws[f"K{r}"].value).strip().lower()
                wpd = str(ws[f"I{r}"].value).strip()
                name = str(ws[f"D{r}"].value or "").strip()

                bad = False
                if nat == "Singapore" and pr == "pr": bad = True
                if idt != "NRIC" and pr == "pr": bad = True
                if idt == "FIN" and (nat == "Singapore" or pr == "pr"): bad = True
                if idt == "NRIC" and not (nat == "Singapore" or pr == "pr"): bad = True
                if idt == "FIN" and not wpd: bad = True
                if idt == "WP" and not wpd: bad = True

                if bad:
                    for col in ("G", "J", "K", "I"):
                        ws[f"{col}{r}"].fill = warning_fill
                    errors += 1

                if name:
                    if name in seen:
                        ws[f"D{r}"].fill = warning_fill
                        ws[f"D{seen[name]}"] .fill = warning_fill
                        errors += 1
                    else:
                        seen[name] = r

            if errors:
                st.warning(f"⚠️ {errors} validation error(s) found.")

            column_widths = {
                "A": 3.38, "C": 23.06, "D": 17.25, "E": 17.63, "F": 26.25,
                "G": 13.94, "H": 24.06, "I": 18.38, "J": 20.31, "K": 4,
                "L": 5.81, "M": 11.5,
            }

            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                if col_letter == "B":
                    width = max(len(str(cell.value)) for cell in col if cell.value)
                    ws.column_dimensions[col_letter].width = width
                elif col_letter in column_widths:
                    ws.column_dimensions[col_letter].width = column_widths[col_letter]

            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 16.8

            plates = []
            for v in df["Vehicle Plate Number"].dropna():
                plates += [x.strip() for x in str(v).split(";") if x.strip()]
            ins = ws.max_row + 2
            if plates:
                ws[f"B{ins}"].value = "Vehicles"
                ws[f"B{ins}"].font = Font(size=9)
                ws[f"B{ins}"].border = border
                ws[f"B{ins}"].alignment = center

                ws[f"B{ins+1}"].value = ";".join(sorted(set(plates)))
                ws[f"B{ins+1}"].font = Font(size=9)
                ws[f"B{ins+1}"].border = border
                ws[f"B{ins+1}"].alignment = center
                ins += 2

            ws[f"B{ins}"].value = "Total Visitors"
            ws[f"B{ins}"].font = Font(size=9)
            ws[f"B{ins}"].border = border
            ws[f"B{ins}"].alignment = center

            ws[f"B{ins+1}"].value = df["Company Full Name"].notna().sum()
            ws[f"B{ins+1}"].font = Font(size=9)
            ws[f"B{ins+1}"].border = border
            ws[f"B{ins+1}"].alignment = center

    buf.seek(0)
    return buf

# --- Main Execution ---
if uploaded:
    all_sheets = pd.read_excel(uploaded, sheet_name=None)
    raw_df = all_sheets.get("Visitor List")

    if raw_df is not None:
        company_cell = raw_df.iloc[0, 2]
        company = (
            str(company_cell).strip()
            if pd.notna(company_cell) and str(company_cell).strip()
            else "VisitorList"
        )

        cleaned = clean_data(raw_df)
        all_sheets["Visitor List"] = cleaned
        out_buf = generate_full_workbook(all_sheets)

        today_str = datetime.now(ZoneInfo("Asia/Singapore")).strftime("%Y%m%d")
        fname = f"{company}_{today_str}.xlsx"

        st.download_button(
            label="📅 Download Cleaned Workbook",
            data=out_buf.getvalue(),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption(
            "✅ Your data has been validated. Please double-check critical fields before sharing with DC team."
        )
    else:
        st.error("❌ 'Visitor List' tab not found in uploaded file.")
